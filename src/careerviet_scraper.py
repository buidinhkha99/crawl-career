#!/usr/bin/env python3
"""
CareerViet Scraper - Multi-thread, fetch all pages
Author: AI Assistant
Date: 2026-03-17

Phase 1: Fetch tất cả listing pages song song → lấy URLs + basic info
Phase 2: Fetch tất cả detail pages song song → lấy Ngành nghề, Kinh nghiệm, Ngày hết hạn

Yêu cầu:
    pip install requests pandas openpyxl beautifulsoup4 lxml

Usage:
    python careerviet_scraper.py --workers 15 --output careerviet_jobs.xlsx
    python careerviet_scraper.py --pages 50 --workers 15   # giới hạn số trang listing
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from bs4 import BeautifulSoup
import argparse
import sys
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin
import threading


BASE_URL = "https://careerviet.vn"
LIST_URL_PAGE1 = "https://careerviet.vn/viec-lam/tat-ca-viec-lam-vi.html"
LIST_URL_PAGEN = "https://careerviet.vn/viec-lam/tat-ca-viec-lam-trang-{page}-vi.html"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8',
    'Referer': 'https://careerviet.vn/',
}


# ─────────────────────────────────────────────
# PHASE 1: Listing pages
# ─────────────────────────────────────────────

def get_total_pages(session):
    """Lấy tổng số trang từ trang 1."""
    try:
        resp = session.get(LIST_URL_PAGE1, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')

        # Tìm số trang cuối trong pagination
        pag = soup.find('div', class_='pagination')
        if pag:
            page_links = pag.find_all('a', href=re.compile(r'trang-(\d+)'))
            if page_links:
                nums = [int(re.search(r'trang-(\d+)', a['href']).group(1)) for a in page_links]
                return max(nums)
            # Tìm nút "..." hoặc số cuối
            all_nums = re.findall(r'trang-(\d+)', str(pag))
            if all_nums:
                return max(int(n) for n in all_nums)

        # Fallback: tính từ tổng số jobs
        text = soup.get_text()
        total_match = re.search(r'([\d,\.]+)\s*việc làm', text, re.I)
        if total_match:
            total = int(total_match.group(1).replace(',', '').replace('.', ''))
            jobs_per_page = len(parse_listing_page(soup))
            if jobs_per_page > 0:
                return (total // jobs_per_page) + 1

        return 1  # fallback
    except Exception as e:
        print(f"    ⚠ Không xác định được tổng trang: {e}")
        return 1


def parse_listing_page(soup):
    """Parse HTML listing page → list of job dicts (basic info)."""
    jobs = []
    job_items = soup.find_all('div', class_=re.compile(r'job-item'))
    if not job_items:
        job_items = soup.find_all(attrs={'data-job-id': True})

    ngay_cap_nhat = datetime.now().strftime('%d/%m/%Y')

    for item in job_items:
        try:
            # Tiêu đề + Link
            link_elem = item.find('a', class_='job_link')
            if not link_elem:
                link_elem = item.find('a', href=re.compile(r'/viec-lam/'))
            title = link_elem.get_text(strip=True) if link_elem else 'N/A'
            href = link_elem.get('href', '') if link_elem else ''
            job_link = href if href.startswith('http') else urljoin(BASE_URL, href) if href else 'N/A'

            # Công ty
            company_elem = item.find(class_=re.compile(r'company|employer', re.I))
            company = company_elem.get_text(strip=True) if company_elem else 'N/A'

            # Mức lương
            salary_elem = item.find(class_=re.compile(r'salary|wage', re.I))
            salary = salary_elem.get_text(strip=True) if salary_elem else 'N/A'

            # Địa điểm
            loc_elem = item.find(class_=re.compile(r'location|address|city', re.I))
            location = loc_elem.get_text(strip=True) if loc_elem else 'N/A'

            # Ngày đăng
            date_elem = item.find('time') or item.find(class_=re.compile(r'date|posted', re.I))
            posted_date = date_elem.get_text(strip=True) if date_elem else 'N/A'

            if title != 'N/A' and job_link != 'N/A':
                jobs.append({
                    'Tiêu đề công việc': title,
                    'Công ty': company,
                    'Mức lương': salary,
                    'Địa điểm': location,
                    'Ngày đăng': posted_date,
                    'Ngày cập nhật': ngay_cap_nhat,
                    'Link': job_link,
                    # detail fields — sẽ fill ở phase 2
                    'Ngày hết hạn': 'N/A',
                    'Ngành hoạt động': 'N/A',
                    'Kinh nghiệm': 'N/A',
                })
        except Exception:
            continue
    return jobs


def fetch_listing_page(page_num, session, print_lock, counter, total_pages):
    """Fetch một trang listing, trả về list job dicts."""
    url = LIST_URL_PAGE1 if page_num == 1 else LIST_URL_PAGEN.format(page=page_num)
    try:
        resp = session.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')
        jobs = parse_listing_page(soup)
        with print_lock:
            counter[0] += 1
            print(f"  Listing [{counter[0]}/{total_pages}] trang {page_num}: {len(jobs)} jobs", flush=True)
        return jobs
    except Exception as e:
        with print_lock:
            counter[0] += 1
            print(f"  Listing [{counter[0]}/{total_pages}] trang {page_num}: LỖI — {e}", flush=True)
        return []


# ─────────────────────────────────────────────
# PHASE 2: Detail pages
# ─────────────────────────────────────────────

def fetch_job_detail(job, session, print_lock, counter, total):
    """Fetch trang chi tiết job, bổ sung Ngành nghề / Kinh nghiệm / Ngày hết hạn."""
    url = job['Link']
    try:
        resp = session.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.content, 'html.parser')
        text = soup.get_text()

        # Ngày hết hạn
        dl_match = re.search(r'Hết hạn nộp[:\s]*(\d{2}/\d{2}/\d{4})', text)
        if dl_match:
            job['Ngày hết hạn'] = dl_match.group(1)
        else:
            t = soup.find('time', {'itemprop': 'validThrough'})
            if t:
                job['Ngày hết hạn'] = t.get('datetime', 'N/A')

        # Ngành nghề — lấy occurrence cuối (trong section job detail)
        positions = [m.start() for m in re.finditer(r'Ngành nghề', text)]
        if positions:
            context = text[positions[-1]:positions[-1] + 300]
            m = re.search(
                r'Ngành nghề\s*(.+?)(?:Hình thức|Lương|Kinh nghiệm|Cấp bậc|Hết hạn|$)',
                context, re.I | re.DOTALL
            )
            if m:
                val = re.sub(r'\s+', ' ', m.group(1)).strip()
                if 3 < len(val) < 200:
                    job['Ngành hoạt động'] = val

        # Kinh nghiệm
        exp_m = re.search(
            r'Kinh nghiệm[:\s]*(.+?)(?:Cấp bậc|Hết hạn|Hình thức|Phúc lợi|$)',
            text, re.I | re.DOTALL
        )
        if exp_m:
            val = re.sub(r'\s+', ' ', exp_m.group(1)).strip()
            if val and len(val) < 80:
                job['Kinh nghiệm'] = val

        with print_lock:
            counter[0] += 1
            done = counter[0]
            pct = done * 100 // total
            print(f"  Detail [{done}/{total} {pct}%] ✓ {job['Tiêu đề công việc'][:50]}", flush=True)

    except Exception as e:
        with print_lock:
            counter[0] += 1
            done = counter[0]
            pct = done * 100 // total
            print(f"  Detail [{done}/{total} {pct}%] ✗ {url.split('/')[-1][:40]} — {e}", flush=True)

    return job


# ─────────────────────────────────────────────
# MAIN SCRAPE
# ─────────────────────────────────────────────

def scrape_careerviet(pages=None, workers=15):
    """
    Scrape CareerViet với 2 phase song song.

    Args:
        pages: Số trang listing cần lấy (None = tất cả)
        workers: Số luồng song song

    Returns:
        DataFrame chứa thông tin công việc
    """
    session = requests.Session()
    print_lock = threading.Lock()

    # Xác định tổng số trang
    print("[→] Đang xác định tổng số trang...")
    total_pages = get_total_pages(session)
    if pages:
        total_pages = min(pages, total_pages)
    print(f"    ✓ Sẽ fetch {total_pages} trang listing\n")

    # ── PHASE 1: Fetch listing pages song song ──
    print(f"[PHASE 1] Fetch listing pages với {workers} workers...")
    print("=" * 60)

    counter = [0]
    all_jobs = []

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(fetch_listing_page, p, session, print_lock, counter, total_pages): p
            for p in range(1, total_pages + 1)
        }
        for future in as_completed(futures):
            all_jobs.extend(future.result())

    # Dedup theo link
    seen = set()
    unique_jobs = []
    for job in all_jobs:
        if job['Link'] not in seen:
            seen.add(job['Link'])
            unique_jobs.append(job)

    print(f"\n    ✓ Tổng: {len(unique_jobs)} jobs unique từ {total_pages} trang")

    if not unique_jobs:
        return pd.DataFrame()

    # ── PHASE 2: Fetch detail pages song song ──
    print(f"\n[PHASE 2] Fetch detail pages với {workers} workers...")
    print("=" * 60)

    counter = [0]
    total_jobs = len(unique_jobs)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(fetch_job_detail, job, session, print_lock, counter, total_jobs): i
            for i, job in enumerate(unique_jobs)
        }
        results = []
        for future in as_completed(futures):
            results.append(future.result())

    print("=" * 60)
    print(f"✓ Hoàn thành {len(results)} jobs")
    return pd.DataFrame(results)


# ─────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────

def create_excel_report(df, output_path):
    """Tạo báo cáo Excel từ DataFrame."""
    wb = Workbook()
    header_dark = "333333"
    blue_primary = "0066CC"
    link_color = "0563C1"
    header_fill = PatternFill(start_color=header_dark, end_color=header_dark, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # === COVER ===
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_view.showGridLines = False

    ws_cover.merge_cells('B2:G2')
    ws_cover['B2'] = "BÁO CÁO THÔNG TIN TUYỂN DỤNG CAREERVIET"
    ws_cover['B2'].font = Font(size=20, bold=True, color=header_dark)
    ws_cover['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cover.row_dimensions[2].height = 35

    ws_cover.merge_cells('B4:G4')
    ws_cover['B4'] = "Dữ liệu được thu thập từ CareerViet.vn"
    ws_cover['B4'].font = Font(size=12, color="666666")
    ws_cover['B4'].alignment = Alignment(horizontal='center')

    ws_cover['B7'] = "TỔNG QUAN"
    ws_cover['B7'].font = Font(size=14, bold=True, color=blue_primary)

    metrics = [
        ("Tổng số công việc", len(df)),
        ("Số công ty", df['Công ty'].nunique() if len(df) > 0 else 0),
        ("Số địa điểm", df['Địa điểm'].nunique() if len(df) > 0 else 0),
        ("Ngày cập nhật", df['Ngày cập nhật'].iloc[0] if len(df) > 0 else 'N/A'),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 9 + i
        ws_cover[f'B{row}'] = label
        ws_cover[f'B{row}'].font = Font(color="666666")
        ws_cover[f'D{row}'] = value
        ws_cover[f'D{row}'].font = Font(bold=True, color=header_dark)

    ws_cover.column_dimensions['A'].width = 3
    ws_cover.column_dimensions['B'].width = 25
    ws_cover.column_dimensions['C'].width = 5
    ws_cover.column_dimensions['D'].width = 35

    # === DANH SÁCH CÔNG VIỆC ===
    ws_jobs = wb.create_sheet("Danh sách công việc")
    ws_jobs.sheet_view.showGridLines = False

    col_headers = ['STT', 'Tiêu đề công việc', 'Công ty', 'Mức lương', 'Địa điểm',
                   'Ngày đăng', 'Ngày hết hạn', 'Ngày cập nhật',
                   'Ngành hoạt động', 'Kinh nghiệm', 'Link']

    for col, header in enumerate(col_headers, 1):
        cell = ws_jobs.cell(row=2, column=col + 1, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws_jobs.row_dimensions[2].height = 30

    col_keys = ['Tiêu đề công việc', 'Công ty', 'Mức lương', 'Địa điểm',
                'Ngày đăng', 'Ngày hết hạn', 'Ngày cập nhật',
                'Ngành hoạt động', 'Kinh nghiệm', 'Link']

    for idx, (_, row_data) in enumerate(df.iterrows()):
        row_num = idx + 3
        ws_jobs.cell(row=row_num, column=2, value=idx + 1).alignment = Alignment(horizontal='center')
        for col_offset, key in enumerate(col_keys, 3):
            value = row_data.get(key, '')
            if key == 'Link' and value and value != 'N/A':
                cell = ws_jobs.cell(row=row_num, column=col_offset)
                cell.value = f'=HYPERLINK("{value}","Xem chi tiết")'
                cell.font = Font(color=link_color, underline='single')
            else:
                ws_jobs.cell(row=row_num, column=col_offset, value=value)

        if idx % 2 == 1:
            for col in range(2, len(col_headers) + 3):
                ws_jobs.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    col_widths = [3, 6, 40, 35, 15, 18, 14, 14, 14, 30, 18, 18]
    for i, w in enumerate(col_widths, ord('A')):
        ws_jobs.column_dimensions[chr(i)].width = w

    # === THỐNG KÊ ===
    ws_stats = wb.create_sheet("Thống kê")
    ws_stats.sheet_view.showGridLines = False
    ws_stats.merge_cells('B2:E2')
    ws_stats['B2'] = "THỐNG KÊ DỮ LIỆU"
    ws_stats['B2'].font = Font(size=16, bold=True, color=header_dark)
    ws_stats['B2'].alignment = Alignment(horizontal='center')
    ws_stats.row_dimensions[2].height = 30

    if len(df) > 0:
        def write_stat(ws, start_row, title, series):
            ws[f'B{start_row}'] = title
            ws[f'B{start_row}'].font = Font(size=12, bold=True, color=blue_primary)
            hr = start_row + 2
            for col, h in enumerate(['Giá trị', 'Số lượng'], 1):
                cell = ws.cell(row=hr, column=col + 1, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            stats = series.value_counts().head(20)
            for i, (val, cnt) in enumerate(stats.items()):
                ws.cell(row=hr + 1 + i, column=2, value=str(val))
                ws.cell(row=hr + 1 + i, column=3, value=cnt).alignment = Alignment(horizontal='center')
            return hr + 1 + len(stats) + 2

        next_row = write_stat(ws_stats, 5, "Thống kê theo địa điểm", df['Địa điểm'])
        next_row = write_stat(ws_stats, next_row, "Thống kê theo ngành hoạt động", df['Ngành hoạt động'])
        write_stat(ws_stats, next_row, "Thống kê theo mức lương", df['Mức lương'])

    ws_stats.column_dimensions['A'].width = 3
    ws_stats.column_dimensions['B'].width = 40
    ws_stats.column_dimensions['C'].width = 15

    wb.save(output_path)
    print(f"✅ Đã lưu: {output_path}")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='CareerViet Scraper - Multi-thread, all pages')
    parser.add_argument('--pages', type=int, default=None,
                        help='Số trang listing cần lấy (mặc định: TẤT CẢ)')
    parser.add_argument('--workers', type=int, default=15,
                        help='Số luồng song song (mặc định: 15)')
    parser.add_argument('--output', type=str, default='careerviet_jobs.xlsx',
                        help='File Excel đầu ra')
    args = parser.parse_args()

    print("=" * 60)
    print("CAREERVIET SCRAPER - Multi-thread")
    print("=" * 60)
    print(f"Trang listing: {args.pages or 'TẤT CẢ'}")
    print(f"Workers: {args.workers}")
    print(f"Output: {args.output}")
    print("=" * 60)

    df = scrape_careerviet(pages=args.pages, workers=args.workers)

    if len(df) == 0:
        print("\n✗ Không lấy được dữ liệu nào!")
        sys.exit(1)

    print(f"\n✓ Tổng cộng: {len(df)} công việc")
    create_excel_report(df, args.output)
    print("\n" + "=" * 60)
    print("HOÀN THÀNH!")
    print("=" * 60)


if __name__ == "__main__":
    main()
