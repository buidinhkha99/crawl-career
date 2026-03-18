#!/usr/bin/env python3
"""
TopCV Scraper API - Sử dụng JSON API của TopCV
Author: AI Assistant
Date: 2026-02-06

Yêu cầu:
    pip install requests pandas openpyxl

Usage:
    python topcv_scraper_api.py --pages 100 --output topcv_jobs.xlsx
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import time
import argparse
import sys
from datetime import datetime
import json


def init_chrome_driver():
    """
    Khởi tạo Chromium driver dùng chung cho toàn bộ phiên scrape
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager

    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--log-level=3')

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_page_load_timeout(20)
    return driver


def get_job_industry_chromium(job_url, driver=None):
    """
    Sử dụng Chromium để lấy dữ liệu Chuyên môn sau khi JavaScript render.
    Nếu truyền driver thì dùng lại (nhanh hơn), nếu không sẽ tự tạo và đóng.
    """
    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from bs4 import BeautifulSoup
        import re

        own_driver = driver is None
        if own_driver:
            driver = init_chrome_driver()

        soup = None
        try:
            driver.get(job_url)
            # Chờ text "Chuyên môn:" xuất hiện trong page source
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: 'Chuyên môn:' in d.page_source
                )
            except Exception:
                # Timeout (page không có Chuyên môn hoặc render chậm): sleep thêm
                try:
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                except Exception:
                    pass
                time.sleep(4)

            soup = BeautifulSoup(driver.page_source, 'html.parser')
        except Exception:
            # Page load timeout — thử lấy source nếu có thể
            try:
                soup = BeautifulSoup(driver.page_source, 'html.parser')
            except Exception:
                soup = None
        finally:
            if own_driver:
                driver.quit()

        if not soup:
            return None

        # Tìm section "Chuyên môn:"
        for elem in soup.find_all(string=re.compile(r'Chuyên môn:', re.I)):
            parent = elem.parent
            if parent and parent.get('class') and 'job-tags__group-name' in ' '.join(parent.get('class', [])):
                grandparent = parent.parent
                if grandparent:
                    tag_elements = grandparent.find_all(
                        'a', class_=lambda x: x and 'search-from-tag' in (x if isinstance(x, str) else ' '.join(x))
                    )
                    skill_tags = [t.get_text(strip=True) for t in tag_elements
                                  if t.get_text(strip=True) and 1 < len(t.get_text(strip=True)) < 100]
                    if skill_tags:
                        result = ', '.join(skill_tags[:10])
                        if len(result) > 5:
                            return result

        return None
    except Exception:
        return None


def get_job_industry(job_url, session, headers, driver_holder=None):
    """
    Lấy ngành nghề từ trang chi tiết công việc (TopCV).
    - Thử HTTP + parse HTML trước (nhanh)
    - Nếu HTTP fail hoặc không có data → dùng Chromium render JS (chậm hơn nhưng chắc)

    driver_holder: dict {'driver': None} — lazy-init Chromium driver, dùng chung giữa các lần gọi.
    """
    from bs4 import BeautifulSoup
    import re

    # --- Bước 1: Lấy HTML qua HTTP request ---
    soup = None
    try:
        for attempt in range(2):
            try:
                response = session.get(job_url, headers=headers, timeout=20)
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
                break
            except requests.exceptions.Timeout:
                if attempt == 1:
                    break
                time.sleep(1)
            except Exception:
                break
    except Exception:
        pass

    # --- Bước 2: Parse HTML nếu lấy được ---
    if soup:
        # Method 1: Tìm section "Chuyên môn:" trong job-tags
        try:
            for elem in soup.find_all(string=re.compile(r'Chuyên môn:', re.I)):
                parent = elem.parent
                if parent and parent.get('class') and 'job-tags__group-name' in ' '.join(parent.get('class', [])):
                    grandparent = parent.parent
                    if grandparent:
                        # Tags có class "search-from-tag"
                        tag_elements = grandparent.find_all(
                            'a', class_=lambda x: x and 'search-from-tag' in (x if isinstance(x, str) else ' '.join(x))
                        )
                        if tag_elements:
                            tags = [t.get_text(strip=True) for t in tag_elements if t.get_text(strip=True)]
                            if tags:
                                result = ', '.join(tags[:10])
                                if 3 < len(result) < 500:
                                    return result
                        else:
                            # Dùng get_text với separator để tự động thêm dấu phẩy
                            full_text = grandparent.get_text(separator=', ', strip=True)
                            full_text = full_text.replace('Chuyên môn:', '').strip().strip(',').strip()
                            if 3 < len(full_text) < 500:
                                return full_text
        except Exception:
            pass

        # Method 2: JSON-LD schema (industry > skills > occupationalCategory)
        try:
            found_industry = found_skills = found_occ_category = None
            job_levels = {'Nhân viên', 'Quản lý', 'Giám đốc', 'Trưởng phòng', 'Thực tập sinh', 'Mới tốt nghiệp'}

            for script in soup.find_all('script', type='application/ld+json'):
                try:
                    data = json.loads(script.string)
                    if not isinstance(data, dict):
                        continue

                    if not found_industry:
                        ind = data.get('industry', '')
                        if isinstance(ind, str) and len(ind) > 3:
                            found_industry = ind

                    if not found_skills:
                        skills = data.get('skills')
                        if isinstance(skills, str) and 5 < len(skills) < 500:
                            found_skills = skills
                        elif isinstance(skills, list):
                            s = ', '.join(str(x) for x in skills[:8] if x)
                            if len(s) > 3:
                                found_skills = s

                    if not found_occ_category:
                        occ = data.get('occupationalCategory', '')
                        if isinstance(occ, str) and occ not in job_levels and len(occ) > 3:
                            found_occ_category = occ
                except Exception:
                    continue

            if found_industry:
                return found_industry
            if found_skills:
                return found_skills
            if found_occ_category:
                return found_occ_category
        except Exception:
            pass

        # Method 3: Text search "Lĩnh vực:" / "Chuyên môn:"
        try:
            page_text = soup.get_text()
            m = re.search(r'Lĩnh vực:\s*([^\n]+)', page_text)
            if m:
                v = m.group(1).strip()
                if 3 < len(v) < 200:
                    return v
            m = re.search(r'Chuyên môn[:\s]*([^\n]+?)(?:Kỹ năng|Kinh nghiệm|Yêu cầu|$)', page_text)
            if m:
                v = re.sub(r'\s+', ' ', m.group(1).strip())
                if 3 < len(v) < 200:
                    return v
        except Exception:
            pass

    # --- Bước 3: Chromium fallback ---
    driver = driver_holder.get('driver') if driver_holder else None
    print(f"    [Chromium] {job_url.split('/')[-2][:40]}...", flush=True)
    chromium_result = get_job_industry_chromium(job_url, driver=driver)
    return chromium_result or 'N/A'


def scrape_topcv_jobs_api(pages=10, delay=2, get_industry=True):
    """
    Scrape thông tin việc làm từ TopCV sử dụng JSON API

    Args:
        pages: Số trang cần scrape
        delay: Thời gian chờ giữa các request (giây)
        get_industry: Có lấy thông tin ngành nghề từ trang chi tiết không (chậm hơn)

    Returns:
        DataFrame chứa thông tin công việc
    """
    jobs_data = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
        'Referer': 'https://www.topcv.vn/tim-viec-lam-moi-nhat',
        'X-Requested-With': 'XMLHttpRequest'
    }

    session = requests.Session()
    ngay_cap_nhat = datetime.now().strftime('%d/%m/%Y')
    api_url = "https://www.topcv.vn/ajax/viec-lam-moi-nhat"

    # Khởi tạo Chromium driver dùng chung — pre-init trước khi bắt đầu scrape
    driver_holder = None
    if get_industry:
        driver_holder = {'driver': None}
        try:
            print("[Chromium] Khởi tạo trình duyệt dự phòng...")
            driver_holder['driver'] = init_chrome_driver()
            print("[Chromium] Sẵn sàng.")
        except Exception as e:
            print(f"    ⚠ Không thể khởi tạo Chromium: {e}. Sẽ bỏ qua fallback.")

    try:
        for page in range(1, pages + 1):
            params = {'page': page, 'type_keyword': 0, 'sba': 1}
            trang_display = f"{page}/?" if pages >= 99999 else f"{page}/{pages}"
            print(f"[→] Đang scrape trang {trang_display}...")

            try:
                response = session.get(api_url, headers=headers, params=params, timeout=30)
                response.raise_for_status()
                data = response.json()

                if data['status'] != 'success':
                    print(f"    ✗ API trả về lỗi: {data.get('message', 'Unknown error')}")
                    continue

                jobs = data.get('data', [])
                if not jobs:
                    print(f"    ℹ Không có công việc nào ở trang {page}")
                    break

                print(f"    Tìm thấy {len(jobs)} công việc")

                for job in jobs:
                    try:
                        company_info = job.get('company', {})
                        title = job.get('title', 'N/A')
                        company = company_info.get('name', 'N/A')
                        salary = job.get('salary', 'N/A')

                        location = job.get('short_cities', 'N/A')
                        if not location or location == 'N/A':
                            cities_html = job.get('cities', '')
                            if cities_html and '<li>' in cities_html:
                                import re
                                cities_match = re.findall(r'<li>(.*?)<\/li>', cities_html)
                                if cities_match:
                                    location = ', '.join(cities_match)

                        posted_date = job.get('updated_at_str', 'N/A')
                        deadline = job.get('deadline', 'N/A')
                        job_link = job.get('url', 'N/A')

                        # Lấy Chuyên môn: thử HTTP trước, fallback Chromium nếu cần
                        industry = 'N/A'
                        if get_industry and job_link != 'N/A':
                            industry = get_job_industry(job_link, session, headers, driver_holder=driver_holder)
                            time.sleep(0.3)

                        jobs_data.append({
                            'Tiêu đề công việc': title,
                            'Công ty': company,
                            'Mức lương': salary,
                            'Địa điểm': location,
                            'Ngày đăng': posted_date,
                            'Ngày hết hạn': deadline,
                            'Ngày cập nhật': ngay_cap_nhat,
                            'Ngành hoạt động': industry,
                            'Link': job_link
                        })
                    except Exception as e:
                        print(f"    ⚠ Lỗi khi xử lý job: {e}")
                        continue

                print(f"    ✓ Đã xử lý {len(jobs)} công việc ở trang {page}")
                time.sleep(delay)

            except requests.exceptions.Timeout:
                print(f"    ✗ Timeout khi tải trang {page}")
                time.sleep(delay * 2)
                continue
            except requests.exceptions.RequestException as e:
                print(f"    ✗ Lỗi kết nối: {e}")
                if '429' in str(e):
                    print(f"    ⏳ Rate limit! Chờ {delay * 3} giây...")
                    time.sleep(delay * 3)
                else:
                    time.sleep(delay)
                continue
            except Exception as e:
                print(f"    ✗ Lỗi không xác định: {e}")
                continue

    finally:
        if driver_holder and driver_holder.get('driver'):
            driver_holder['driver'].quit()
            print("[Chromium] Đã đóng trình duyệt.")

    return pd.DataFrame(jobs_data)


def create_excel_report(df, output_path):
    """
    Tạo báo cáo Excel từ DataFrame với HYPERLINK
    """
    wb = Workbook()

    # Colors
    header_dark = "333333"
    blue_primary = "0066CC"
    link_color = "0563C1"

    # === SHEET 1: COVER ===
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_view.showGridLines = False

    ws_cover.merge_cells('B2:G2')
    ws_cover['B2'] = "BÁO CÁO THÔNG TIN TUYỂN DỤNG TOPCV"
    ws_cover['B2'].font = Font(size=20, bold=True, color=header_dark)
    ws_cover['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cover.row_dimensions[2].height = 35

    ws_cover.merge_cells('B4:G4')
    ws_cover['B4'] = "Dữ liệu được thu thập từ TopCV.vn API"
    ws_cover['B4'].font = Font(size=12, color="666666")
    ws_cover['B4'].alignment = Alignment(horizontal='center')

    ws_cover['B7'] = "TỔNG QUAN"
    ws_cover['B7'].font = Font(size=14, bold=True, color=blue_primary)

    metrics = [
        ("Tổng số công việc", len(df)),
        ("Số công ty", df['Công ty'].nunique() if len(df) > 0 else 0),
        ("Số địa điểm", df['Địa điểm'].nunique() if len(df) > 0 else 0),
        ("Ngày cập nhật", df['Ngày cập nhật'].iloc[0] if len(df) > 0 else 'N/A')
    ]

    for i, (label, value) in enumerate(metrics):
        row = 9 + i
        ws_cover[f'B{row}'] = label
        ws_cover[f'B{row}'].font = Font(color="666666")
        ws_cover[f'D{row}'] = value
        ws_cover[f'D{row}'].font = Font(bold=True, color=header_dark)

    ws_cover.column_dimensions['A'].width = 3
    ws_cover.column_dimensions['B'].width = 25
    ws_cover.column_dimensions['C'].width = 5
    ws_cover.column_dimensions['D'].width = 35

    # === SHEET 2: DANH SÁCH CÔNG VIỆC ===
    ws_jobs = wb.create_sheet("Danh sách công việc")
    ws_jobs.sheet_view.showGridLines = False

    headers = ['STT', 'Tiêu đề công việc', 'Công ty', 'Mức lương', 'Địa điểm',
               'Ngày đăng', 'Ngày hết hạn', 'Ngày cập nhật', 'Ngành hoạt động', 'Link']
    header_fill = PatternFill(start_color=header_dark, end_color=header_dark, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws_jobs.cell(row=2, column=col+1, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws_jobs.row_dimensions[2].height = 30

    # Data với HYPERLINK
    for idx, row_data in df.iterrows():
        row_num = idx + 3
        ws_jobs.cell(row=row_num, column=2, value=idx + 1).alignment = Alignment(horizontal='center')
        ws_jobs.cell(row=row_num, column=3, value=row_data.get('Tiêu đề công việc', ''))
        ws_jobs.cell(row=row_num, column=4, value=row_data.get('Công ty', ''))
        ws_jobs.cell(row=row_num, column=5, value=row_data.get('Mức lương', ''))
        ws_jobs.cell(row=row_num, column=6, value=row_data.get('Địa điểm', ''))
        ws_jobs.cell(row=row_num, column=7, value=row_data.get('Ngày đăng', ''))
        ws_jobs.cell(row=row_num, column=8, value=row_data.get('Ngày hết hạn', ''))
        ws_jobs.cell(row=row_num, column=9, value=row_data.get('Ngày cập nhật', ''))
        ws_jobs.cell(row=row_num, column=10, value=row_data.get('Ngành hoạt động', ''))

        # HYPERLINK cho cột Link
        link = row_data.get('Link', '')
        if link and link != 'N/A':
            link_cell = ws_jobs.cell(row=row_num, column=11)
            link_cell.value = f'=HYPERLINK("{link}","Xem chi tiết")'
            link_cell.font = Font(color=link_color, underline='single')
            link_cell.alignment = Alignment(horizontal='left')
        else:
            ws_jobs.cell(row=row_num, column=11, value='N/A')

        if idx % 2 == 1:
            for col in range(2, 12):
                ws_jobs.cell(row=row_num, column=col).fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Column widths
    ws_jobs.column_dimensions['A'].width = 3
    ws_jobs.column_dimensions['B'].width = 6
    ws_jobs.column_dimensions['C'].width = 40
    ws_jobs.column_dimensions['D'].width = 35
    ws_jobs.column_dimensions['E'].width = 15
    ws_jobs.column_dimensions['F'].width = 18
    ws_jobs.column_dimensions['G'].width = 14
    ws_jobs.column_dimensions['H'].width = 14
    ws_jobs.column_dimensions['I'].width = 14
    ws_jobs.column_dimensions['J'].width = 25
    ws_jobs.column_dimensions['K'].width = 18

    # === SHEET 3: THỐNG KÊ ===
    ws_stats = wb.create_sheet("Thống kê")
    ws_stats.sheet_view.showGridLines = False

    ws_stats.merge_cells('B2:E2')
    ws_stats['B2'] = "THỐNG KÊ DỮ LIỆU"
    ws_stats['B2'].font = Font(size=16, bold=True, color=header_dark)
    ws_stats['B2'].alignment = Alignment(horizontal='center')
    ws_stats.row_dimensions[2].height = 30

    if len(df) > 0:
        current_row = 5

        # Thống kê theo địa điểm
        ws_stats[f'B{current_row}'] = "Thống kê theo địa điểm"
        ws_stats[f'B{current_row}'].font = Font(size=12, bold=True, color=blue_primary)

        location_stats = df['Địa điểm'].value_counts().reset_index()
        location_stats.columns = ['Địa điểm', 'Số lượng']

        header_row = current_row + 2
        for col, header in enumerate(['Địa điểm', 'Số lượng'], 1):
            cell = ws_stats.cell(row=header_row, column=col+1, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for idx, row_data in location_stats.head(20).iterrows():
            row_num = header_row + 1 + idx
            ws_stats.cell(row=row_num, column=2, value=row_data['Địa điểm'])
            ws_stats.cell(row=row_num, column=3, value=row_data['Số lượng']).alignment = Alignment(horizontal='center')

        # Thống kê theo lương
        current_row = header_row + min(len(location_stats), 20) + 3
        ws_stats[f'B{current_row}'] = "Thống kê theo mức lương"
        ws_stats[f'B{current_row}'].font = Font(size=12, bold=True, color=blue_primary)

        salary_stats = df['Mức lương'].value_counts().reset_index()
        salary_stats.columns = ['Mức lương', 'Số lượng']

        header_row = current_row + 2
        for col, header in enumerate(['Mức lương', 'Số lượng'], 1):
            cell = ws_stats.cell(row=header_row, column=col+1, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for idx, row_data in salary_stats.head(20).iterrows():
            row_num = header_row + 1 + idx
            ws_stats.cell(row=row_num, column=2, value=row_data['Mức lương'])
            ws_stats.cell(row=row_num, column=3, value=row_data['Số lượng']).alignment = Alignment(horizontal='center')

    ws_stats.column_dimensions['A'].width = 3
    ws_stats.column_dimensions['B'].width = 40
    ws_stats.column_dimensions['C'].width = 15

    # Save
    wb.save(output_path)
    print(f"✅ Đã lưu báo cáo: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='TopCV Scraper API - Lấy thông tin tuyển dụng từ TopCV.vn API')
    parser.add_argument('--pages', type=int, default=10, help='Số trang cần scrape (mặc định: 10)')
    parser.add_argument('--all', action='store_true', help='Crawl toàn bộ dữ liệu đến trang cuối')
    parser.add_argument('--output', type=str, default='topcv_jobs_api.xlsx', help='Tên file Excel đầu ra')
    parser.add_argument('--delay', type=float, default=2, help='Thời gian chờ giữa các request (giây)')
    parser.add_argument('--with-industry', action='store_true', help='Lấy thông tin ngành nghề từ trang chi tiết (mặc định: Có)')
    parser.add_argument('--no-industry', action='store_true', help='KHÔNG lấy ngành nghề (nhanh hơn)')

    args = parser.parse_args()

    # --all: dùng số trang rất lớn, scraper sẽ tự dừng khi hết data
    pages = 99999 if args.all else args.pages
    get_industry = not args.no_industry

    print("=" * 60)
    print("TOPCV SCRAPER API - Sử dụng JSON API")
    print("=" * 60)
    print(f"Chế độ: {'Toàn bộ dữ liệu' if args.all else f'{pages} trang'}")
    print(f"File đầu ra: {args.output}")
    print(f"Delay: {args.delay}s")
    print(f"Lấy ngành nghề: {'Có' if get_industry else 'Không (nhanh hơn)'}")
    print("=" * 60)

    df = scrape_topcv_jobs_api(pages=pages, delay=args.delay, get_industry=get_industry)

    if len(df) == 0:
        print("\n✗ Không lấy được dữ liệu nào!")
        print("Có thể do:")
        print("  - API đang bảo trì")
        print("  - Kết nối mạng không ổn định")
        print("  - API đã thay đổi cấu trúc")
        sys.exit(1)

    print(f"\n✓ Tổng cộng: {len(df)} công việc")

    # Create Excel report
    create_excel_report(df, args.output)

    print("\n" + "=" * 60)
    print("HOÀN THÀNH!")
    print("=" * 60)


if __name__ == "__main__":
    main()
