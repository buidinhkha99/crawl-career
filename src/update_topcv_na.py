#!/usr/bin/env python3
"""
Update NA values in "Ngành hoạt động" column of topcv_full.xlsx
by visiting each job detail page and scraping "Chuyên môn" data.

Based on logic from topcv_scraper_api.py

Usage:
    python src/update_topcv_na.py --input data/topcv_full.xlsx
    python src/update_topcv_na.py --input data/topcv_full.xlsx --output data/topcv_updated.xlsx
"""

import requests
import pandas as pd
import time
import argparse
import json
import re
import sys
from bs4 import BeautifulSoup


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
}


def init_chrome_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager

    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--log-level=3')

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_page_load_timeout(20)
    return driver


def parse_chuyen_mon(soup):
    """
    Parse "Chuyên môn" tags from a BeautifulSoup object.
    Returns a comma-separated string or None.
    """
    # Method 1: Find section "Chuyên môn:" in job-tags
    for elem in soup.find_all(string=re.compile(r'Chuyên môn:', re.I)):
        parent = elem.parent
        if parent and parent.get('class') and 'job-tags__group-name' in ' '.join(parent.get('class', [])):
            grandparent = parent.parent
            if grandparent:
                tag_elements = grandparent.find_all(
                    'a', class_=lambda x: x and 'search-from-tag' in (x if isinstance(x, str) else ' '.join(x))
                )
                if tag_elements:
                    tags = [t.get_text(strip=True) for t in tag_elements if t.get_text(strip=True)]
                    if tags:
                        result = ', '.join(tags[:10])
                        if 3 < len(result) < 500:
                            return result
                else:
                    full_text = grandparent.get_text(separator=', ', strip=True)
                    full_text = full_text.replace('Chuyên môn:', '').strip().strip(',').strip()
                    if 3 < len(full_text) < 500:
                        return full_text

    # Method 2: JSON-LD schema
    job_levels = {'Nhân viên', 'Quản lý', 'Giám đốc', 'Trưởng phòng', 'Thực tập sinh', 'Mới tốt nghiệp'}
    found_industry = found_skills = found_occ_category = None

    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string)
            if not isinstance(data, dict):
                continue

            if not found_industry:
                ind = data.get('industry', '')
                if isinstance(ind, str) and len(ind) > 3:
                    found_industry = ind

            if not found_skills:
                skills = data.get('skills')
                if isinstance(skills, str) and 5 < len(skills) < 500:
                    found_skills = skills
                elif isinstance(skills, list):
                    s = ', '.join(str(x) for x in skills[:8] if x)
                    if len(s) > 3:
                        found_skills = s

            if not found_occ_category:
                occ = data.get('occupationalCategory', '')
                if isinstance(occ, str) and occ not in job_levels and len(occ) > 3:
                    found_occ_category = occ
        except Exception:
            continue

    if found_industry:
        return found_industry
    if found_skills:
        return found_skills
    if found_occ_category:
        return found_occ_category

    # Method 3: Text search
    page_text = soup.get_text()
    m = re.search(r'Lĩnh vực:\s*([^\n]+)', page_text)
    if m:
        v = m.group(1).strip()
        if 3 < len(v) < 200:
            return v
    m = re.search(r'Chuyên môn[:\s]*([^\n]+?)(?:Kỹ năng|Kinh nghiệm|Yêu cầu|$)', page_text)
    if m:
        v = re.sub(r'\s+', ' ', m.group(1).strip())
        if 3 < len(v) < 200:
            return v

    return None


def get_chuyen_mon_chromium(job_url, driver):
    """
    Use Chromium to render the page and extract "Chuyên môn" data.
    """
    from selenium.webdriver.support.ui import WebDriverWait

    try:
        driver.get(job_url)
        try:
            WebDriverWait(driver, 15).until(
                lambda d: 'Chuyên môn:' in d.page_source
            )
        except Exception:
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: 'job-tags' in d.page_source
                )
            except Exception:
                time.sleep(4)

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        return parse_chuyen_mon(soup)
    except Exception as e:
        print(f"      [Chromium error] {e}")
        return None


def get_chuyen_mon(job_url, session, driver):
    """
    Try HTTP parse first, fallback to Chromium.
    """
    # Step 1: HTTP request
    soup = None
    try:
        for attempt in range(2):
            try:
                response = session.get(job_url, headers=HEADERS, timeout=20)
                response.raise_for_status()
                soup = BeautifulSoup(response.content, 'html.parser')
                break
            except requests.exceptions.Timeout:
                if attempt == 1:
                    break
                time.sleep(1)
            except Exception:
                break
    except Exception:
        pass

    # Step 2: Parse HTML
    if soup:
        result = parse_chuyen_mon(soup)
        if result:
            return result

    # Step 3: Chromium fallback
    if driver:
        print(f"      [Chromium] {job_url.split('/')[-2][:50]}", flush=True)
        result = get_chuyen_mon_chromium(job_url, driver)
        if result:
            return result

    return 'N/A'


def is_na(value):
    """Check if a cell value is considered NA."""
    if value is None:
        return True
    s = str(value).strip()
    return s == '' or s == 'N/A' or s.lower() == 'nan' or s.lower() == 'none'


def load_na_rows_from_excel(input_path):
    """
    Use openpyxl to read NA rows directly, extracting URLs from HYPERLINK formulas.
    Returns: (workbook, worksheet, nganh_col, list of (excel_row, title, url))
    """
    from openpyxl import load_workbook
    wb = load_workbook(input_path)

    # Find jobs sheet
    ws = None
    for name in wb.sheetnames:
        if 'danh sách' in name.lower() or 'công việc' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Find column indices from row 2 (header row)
    nganh_col = title_col = link_col = None
    for cell in ws[2]:
        if cell.value:
            v = str(cell.value).lower().strip()
            if 'ngành hoạt động' in v:
                nganh_col = cell.column
            elif 'tiêu đề' in v or 'title' in v:
                title_col = cell.column
            elif v == 'link':
                link_col = cell.column

    print(f"  Sheet: '{ws.title}'")
    print(f"  Columns — Ngành hoạt động: {nganh_col}, Title: {title_col}, Link: {link_col}")

    if not nganh_col or not link_col:
        return wb, ws, nganh_col, []

    na_rows = []
    for row in ws.iter_rows(min_row=3):
        excel_row = row[0].row
        nganh_cell = row[nganh_col - 1]
        link_cell = row[link_col - 1]
        title_cell = row[title_col - 1] if title_col else None

        nganh_val = nganh_cell.value
        if not is_na(nganh_val):
            continue  # already has data

        # Extract URL from HYPERLINK formula
        link_val = link_cell.value
        if not link_val:
            # Try hyperlink attribute
            if link_cell.hyperlink:
                url = link_cell.hyperlink.target
            else:
                continue
        else:
            m = re.match(r'=HYPERLINK\("([^"]+)"', str(link_val), re.I)
            url = m.group(1) if m else str(link_val).strip()

        if not url or url == 'N/A':
            continue

        title = title_cell.value if title_cell else ''
        na_rows.append((excel_row, str(title or ''), url))

    return wb, ws, nganh_col, na_rows


def main():
    parser = argparse.ArgumentParser(description='Update NA Ngành hoạt động in topcv Excel file')
    parser.add_argument('--input', type=str, default='data/topcv_full.xlsx', help='Input Excel file')
    parser.add_argument('--output', type=str, default=None, help='Output file (default: overwrite input)')
    parser.add_argument('--delay', type=float, default=0.5, help='Delay between requests (seconds)')
    parser.add_argument('--no-chromium', action='store_true', help='Disable Chromium fallback')
    parser.add_argument('--limit', type=int, default=0, help='Limit number of rows to process (0 = all)')
    args = parser.parse_args()

    output_path = args.output or args.input

    print("=" * 60)
    print("UPDATE TOPCV NA - Ngành hoạt động")
    print("=" * 60)
    print(f"Input:  {args.input}")
    print(f"Output: {output_path}")
    print(f"Delay:  {args.delay}s")
    print("=" * 60)

    print("\nReading Excel file...")
    wb, ws, nganh_col, na_rows = load_na_rows_from_excel(args.input)
    print(f"NA rows found: {len(na_rows)}")

    if not na_rows:
        print("Nothing to update.")
        return

    if args.limit > 0:
        na_rows = na_rows[:args.limit]
        print(f"Limiting to {args.limit} rows")

    # Init Chromium
    driver = None
    if not args.no_chromium:
        try:
            print("\n[Chromium] Initializing browser...")
            driver = init_chrome_driver()
            print("[Chromium] Ready.")
        except Exception as e:
            print(f"  WARNING: Cannot init Chromium: {e}")

    session = requests.Session()
    updated = 0
    failed = 0

    try:
        for i, (excel_row, title, url) in enumerate(na_rows, 1):
            print(f"  [{i}/{len(na_rows)}] {title[:55]} ...", flush=True)

            result = get_chuyen_mon(url, session, driver)

            # Write directly to the worksheet cell
            ws.cell(row=excel_row, column=nganh_col, value=result)

            if result and result != 'N/A':
                print(f"      => {result}")
                updated += 1
            else:
                print(f"      => N/A (not found)")
                failed += 1

            # Save every 50 rows as checkpoint
            if i % 50 == 0:
                wb.save(output_path)
                print(f"  [Checkpoint saved at row {i}]")

            time.sleep(args.delay)

    finally:
        if driver:
            driver.quit()
            print("\n[Chromium] Browser closed.")

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Updated: {updated} / {len(na_rows)}")
    print(f"Still N/A: {failed}")
    print(f"Saved to: {output_path}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
