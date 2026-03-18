#!/usr/bin/env python3
"""
VietnamWorks Scraper - Sitemap + Parallel HTML Scraping
Author: AI Assistant
Date: 2026-03-17

Lấy job URLs từ sitemap, fetch song song với ThreadPoolExecutor,
extract data từ Next.js SSR embedded JSON.

Yêu cầu:
    pip install requests pandas openpyxl

Usage:
    python vietnamworks_scraper.py --limit 100 --workers 15 --output jobs.xlsx
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import argparse
import sys
import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading


SITEMAP_URL = "https://www.vietnamworks.com/sitemap/jobs.xml"
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8',
}


def get_job_urls_from_sitemap(limit=None):
    """Lấy danh sách job URLs từ sitemap XML. limit=None để lấy tất cả."""
    print(f"[→] Đang lấy job URLs từ sitemap...")
    try:
        resp = requests.get(SITEMAP_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
        ns = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
        urls = []
        for url_elem in root.findall('ns:url', ns):
            loc = url_elem.find('ns:loc', ns)
            if loc is not None and loc.text:
                urls.append(loc.text)
            if limit and len(urls) >= limit:
                break
        print(f"    ✓ Lấy được {len(urls)} job URLs")
        return urls
    except Exception as e:
        print(f"    ✗ Lỗi lấy sitemap: {e}")
        return []


def _extract_from_nextjs_html(html):
    """Extract job data từ Next.js SSR embedded JSON trong HTML."""
    # Gộp tất cả chunks từ self.__next_f.push([1,"..."])
    chunks = re.findall(r'self\.__next_f\.push\(\[1,"(.*?)"\]\)', html, re.DOTALL)
    full_text = ''
    for chunk in chunks:
        try:
            # Decode JSON string escaping đúng cách
            full_text += json.loads(f'"{chunk}"')
        except Exception:
            full_text += chunk

    def get(pattern, text=full_text, group=1):
        m = re.search(pattern, text)
        return m.group(group) if m else None

    # Tiêu đề
    title = get(r'"jobTitle":"([^"]+)"') or 'N/A'
    company = get(r'"companyName":"([^"]+)"') or 'N/A'

    # Lương — ưu tiên prettySalary (có sẵn dạng text)
    salary_raw = get(r'"prettySalary":"([^"]+)"') or get(r'"salaryDisplay":"([^"]+)"') or 'N/A'
    # Fix encoding nếu bị double-encode
    try:
        salary = salary_raw.encode('latin-1').decode('utf-8') if salary_raw != 'N/A' else 'N/A'
    except Exception:
        salary = salary_raw

    # Địa điểm — dùng cityNameVI (Vietnamese name)
    loc_names = re.findall(r'"cityNameVI":"([^"]+)"', full_text)
    if not loc_names:
        loc_names = re.findall(r'"cityName":"([^"]+)"', full_text)
    location = ', '.join(dict.fromkeys(loc_names)) if loc_names else 'N/A'

    # Ngành nghề — ghép "parentNameVI > nameVI" (NGÀNH NGHỀ trên UI)
    parents = re.findall(r'"parentNameVI":"([^"]+)"', full_text)
    children = re.findall(r'"nameVI":"([^"]+)"', full_text)
    if parents and children:
        # Pair parent + child, deduplicate
        pairs = list(dict.fromkeys(
            f"{p} > {c}" for p, c in zip(parents, children)
        ))
        industry = ', '.join(pairs)
    elif parents:
        industry = ', '.join(dict.fromkeys(parents))
    else:
        industry = 'N/A'

    # Cấp bậc
    level = get(r'"jobLevelVI":"([^"]+)"') or get(r'"jobLevel":"([^"]+)"') or 'N/A'

    # Kỹ năng
    skill_names = re.findall(r'"skillName":"([^"]+)"', full_text)
    skills = ', '.join(dict.fromkeys(skill_names)) if skill_names else 'N/A'

    # Phúc lợi
    benefit_names = re.findall(r'"benefitName":"([^"]+)"', full_text)
    benefits = ', '.join(dict.fromkeys(benefit_names)) if benefit_names else 'N/A'

    # Ngày
    created = get(r'"createdOn":"([^"T]+)')
    expired = get(r'"expiredOn":"([^"T]+)')
    try:
        posted_date = datetime.strptime(created, '%Y-%m-%d').strftime('%d/%m/%Y') if created else 'N/A'
        deadline = datetime.strptime(expired, '%Y-%m-%d').strftime('%d/%m/%Y') if expired else 'N/A'
    except Exception:
        posted_date = created or 'N/A'
        deadline = expired or 'N/A'

    return {
        'Tiêu đề công việc': title,
        'Công ty': company,
        'Mức lương': salary,
        'Địa điểm': location,
        'Cấp bậc': level,
        'Ngày đăng': posted_date,
        'Ngày hết hạn': deadline,
        'Ngày cập nhật': datetime.now().strftime('%d/%m/%Y'),
        'Ngành hoạt động': industry,
        'Kỹ năng': skills,
        'Phúc lợi': benefits,
    }


def fetch_job(url, session, print_lock, counter, total):
    """Fetch một job URL và trả về dict data."""
    try:
        resp = session.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        job = _extract_from_nextjs_html(resp.text)
        job['Link'] = url
        with print_lock:
            counter[0] += 1
            done = counter[0]
            pct = done * 100 // total
            title_short = job['Tiêu đề công việc'][:50]
            print(f"  [{done}/{total} {pct}%] ✓ {title_short}", flush=True)
        return job
    except Exception as e:
        with print_lock:
            counter[0] += 1
            done = counter[0]
            pct = done * 100 // total
            print(f"  [{done}/{total} {pct}%] ✗ {url.split('/')[-1][:40]} — {e}", flush=True)
        return None


def scrape_vietnamworks_jobs(limit=None, workers=15):
    """
    Scrape việc làm từ VietnamWorks qua sitemap + song song.

    Args:
        limit: Số job cần scrape (None = tất cả)
        workers: Số luồng song song

    Returns:
        DataFrame chứa thông tin công việc
    """
    job_urls = get_job_urls_from_sitemap(limit=limit)
    if not job_urls:
        return pd.DataFrame()

    print(f"\n[→] Fetch {len(job_urls)} job với {workers} luồng song song...")
    print("=" * 60)

    print_lock = threading.Lock()
    counter = [0]
    jobs_data = []

    session = requests.Session()

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(fetch_job, url, session, print_lock, counter, len(job_urls)): url
            for url in job_urls
        }
        for future in as_completed(futures):
            result = future.result()
            if result and result['Tiêu đề công việc'] != 'N/A':
                jobs_data.append(result)

    print("=" * 60)
    print(f"✓ Thu thập được {len(jobs_data)}/{len(job_urls)} công việc")
    return pd.DataFrame(jobs_data)


def create_excel_report(df, output_path):
    """Tạo báo cáo Excel từ DataFrame."""
    wb = Workbook()
    header_dark = "333333"
    blue_primary = "0066CC"
    link_color = "0563C1"
    header_fill = PatternFill(start_color=header_dark, end_color=header_dark, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # === COVER ===
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_view.showGridLines = False

    ws_cover.merge_cells('B2:G2')
    ws_cover['B2'] = "BÁO CÁO THÔNG TIN TUYỂN DỤNG VIETNAMWORKS"
    ws_cover['B2'].font = Font(size=20, bold=True, color=header_dark)
    ws_cover['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cover.row_dimensions[2].height = 35

    ws_cover.merge_cells('B4:G4')
    ws_cover['B4'] = "Dữ liệu được thu thập từ VietnamWorks.com"
    ws_cover['B4'].font = Font(size=12, color="666666")
    ws_cover['B4'].alignment = Alignment(horizontal='center')

    ws_cover['B7'] = "TỔNG QUAN"
    ws_cover['B7'].font = Font(size=14, bold=True, color=blue_primary)

    metrics = [
        ("Tổng số công việc", len(df)),
        ("Số công ty", df['Công ty'].nunique() if len(df) > 0 else 0),
        ("Số địa điểm", df['Địa điểm'].nunique() if len(df) > 0 else 0),
        ("Ngày cập nhật", df['Ngày cập nhật'].iloc[0] if len(df) > 0 else 'N/A'),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 9 + i
        ws_cover[f'B{row}'] = label
        ws_cover[f'B{row}'].font = Font(color="666666")
        ws_cover[f'D{row}'] = value
        ws_cover[f'D{row}'].font = Font(bold=True, color=header_dark)

    ws_cover.column_dimensions['A'].width = 3
    ws_cover.column_dimensions['B'].width = 25
    ws_cover.column_dimensions['C'].width = 5
    ws_cover.column_dimensions['D'].width = 35

    # === DANH SÁCH CÔNG VIỆC ===
    ws_jobs = wb.create_sheet("Danh sách công việc")
    ws_jobs.sheet_view.showGridLines = False

    headers = ['STT', 'Tiêu đề công việc', 'Công ty', 'Mức lương', 'Địa điểm', 'Cấp bậc',
               'Ngày đăng', 'Ngày hết hạn', 'Ngày cập nhật', 'Ngành hoạt động', 'Kỹ năng', 'Phúc lợi', 'Link']

    for col, header in enumerate(headers, 1):
        cell = ws_jobs.cell(row=2, column=col + 1, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws_jobs.row_dimensions[2].height = 30

    col_keys = ['Tiêu đề công việc', 'Công ty', 'Mức lương', 'Địa điểm', 'Cấp bậc',
                'Ngày đăng', 'Ngày hết hạn', 'Ngày cập nhật', 'Ngành hoạt động', 'Kỹ năng', 'Phúc lợi', 'Link']

    for idx, (_, row_data) in enumerate(df.iterrows()):
        row_num = idx + 3
        ws_jobs.cell(row=row_num, column=2, value=idx + 1).alignment = Alignment(horizontal='center')
        for col_offset, key in enumerate(col_keys, 3):
            value = row_data.get(key, '')
            if key == 'Link' and value and value != 'N/A':
                cell = ws_jobs.cell(row=row_num, column=col_offset)
                cell.value = f'=HYPERLINK("{value}","Xem chi tiết")'
                cell.font = Font(color=link_color, underline='single')
            else:
                ws_jobs.cell(row=row_num, column=col_offset, value=value)

        if idx % 2 == 1:
            for col in range(2, 16):
                ws_jobs.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    col_widths = [3, 6, 40, 30, 15, 18, 12, 14, 14, 14, 25, 30, 30, 18]
    for i, w in enumerate(col_widths, ord('A')):
        ws_jobs.column_dimensions[chr(i)].width = w

    # === THỐNG KÊ ===
    ws_stats = wb.create_sheet("Thống kê")
    ws_stats.sheet_view.showGridLines = False
    ws_stats.merge_cells('B2:E2')
    ws_stats['B2'] = "THỐNG KÊ DỮ LIỆU"
    ws_stats['B2'].font = Font(size=16, bold=True, color=header_dark)
    ws_stats['B2'].alignment = Alignment(horizontal='center')
    ws_stats.row_dimensions[2].height = 30

    if len(df) > 0:
        def write_stat_section(ws, start_row, title, series):
            ws[f'B{start_row}'] = title
            ws[f'B{start_row}'].font = Font(size=12, bold=True, color=blue_primary)
            header_r = start_row + 2
            for col, h in enumerate(['Giá trị', 'Số lượng'], 1):
                cell = ws.cell(row=header_r, column=col + 1, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            stats = series.value_counts().head(20)
            for i, (val, cnt) in enumerate(stats.items()):
                ws.cell(row=header_r + 1 + i, column=2, value=str(val))
                ws.cell(row=header_r + 1 + i, column=3, value=cnt).alignment = Alignment(horizontal='center')
            return header_r + 1 + len(stats) + 2

        next_row = write_stat_section(ws_stats, 5, "Thống kê theo địa điểm", df['Địa điểm'])
        next_row = write_stat_section(ws_stats, next_row, "Thống kê theo cấp bậc", df['Cấp bậc'])
        write_stat_section(ws_stats, next_row, "Thống kê theo mức lương", df['Mức lương'])

    ws_stats.column_dimensions['A'].width = 3
    ws_stats.column_dimensions['B'].width = 40
    ws_stats.column_dimensions['C'].width = 15

    wb.save(output_path)
    print(f"✅ Đã lưu: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='VietnamWorks Scraper - Sitemap + Parallel')
    parser.add_argument('--limit', type=int, default=None, help='Số job cần scrape (mặc định: tất cả)')
    parser.add_argument('--workers', type=int, default=15, help='Số luồng song song (mặc định: 15)')
    parser.add_argument('--output', type=str, default='vietnamworks_jobs.xlsx', help='File Excel đầu ra')
    args = parser.parse_args()

    print("=" * 60)
    print("VIETNAMWORKS SCRAPER - Sitemap + Parallel")
    print("=" * 60)
    print(f"Số job: {args.limit or 'TẤT CẢ'}")
    print(f"Workers: {args.workers}")
    print(f"Output: {args.output}")
    print("=" * 60)

    df = scrape_vietnamworks_jobs(limit=args.limit, workers=args.workers)

    if len(df) == 0:
        print("\n✗ Không lấy được dữ liệu nào!")
        sys.exit(1)

    print(f"\n✓ Tổng cộng: {len(df)} công việc")
    create_excel_report(df, args.output)
    print("\n" + "=" * 60)
    print("HOÀN THÀNH!")
    print("=" * 60)


if __name__ == "__main__":
    main()
